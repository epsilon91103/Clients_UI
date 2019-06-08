from datetime import datetime

from openpyxl import load_workbook


# расположение Excel файлов
PATH_EXCEL_FILES = 'sources/excel/'
# шаблон для пустых полей
EMPTY_FIELD = 'NULL'


# преобразование даты из строчки в класс datetime
def convert_str_to_date(str_date, mask='%Y-%m-%d %H:%M:%S.%f'):
    """
    Date conversion from string to datetime class
    :param str_date: Дата в виде текста
    :param mask: Формат времени
    :return:
    """
    return datetime.strptime(str_date, mask)


# считывание данных из Excel
def read_excel(name_file, sheet_name, header=True, integer_fields=None, date_fields=None):
    """
    Reading and processing data from Excel
    :param name_file: Название файла
    :param sheet_name: Название листа
    :param header: Наличие заголовка
    :param integer_fields: Номера столбцов для преобразования к int
    :param date_fields: Номера столбцов для преобразования в дату
    :return:
    """

    if integer_fields is None:
        integer_fields = []

    if date_fields is None:
        date_fields = []

    wb = load_workbook(PATH_EXCEL_FILES + name_file)
    sheet = wb[sheet_name]

    count_row, count_col = sheet.max_row, sheet.max_column

    data = []
    name_fields = []
    for row in range(1, count_row + 1):
        if row >= 1 + header:
            item = []
            for column in range(1, count_col + 1):
                field = sheet.cell(row=row, column=column).value
                if field == EMPTY_FIELD:
                    field = None
                if column in integer_fields and field is not None:
                    field = int(field)
                if column in date_fields and field is not None:
                    field = convert_str_to_date(field)
                item.append(field)
            data.append(item)
        else:
            name_fields = [sheet.cell(row=row, column=column).value for column in range(1, count_col + 1)]

    wb.close()

    return name_fields, data
